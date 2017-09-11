# 加载模块
import numpy as np
import dbn
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn import preprocessing
from pandas import DataFrame
import os  # 加载路径
from openpyxl.workbook import Workbook  # 保存excel


# 离差标准化
class DeviationStandardization(object):
    """docstring for lichabio"""

    def __init__(self):
        super(DeviationStandardization, self).__init__()

    def MaxMinNormalization(self, x, naxis=0):  # 离差标准化
        self.maxNp = x.max(axis=naxis).astype('float')
        self.minNp = x.min(axis=naxis).astype('float')
        DeviationNp = (x.astype('float') - self.minNp) / \
            (self.maxNp - self.minNp)
        return DeviationNp

    def fanization(self, y):  # 饭离差标准化
        fanDeviationNp = (y.astype('float') *
                          (self.maxNp - self.minNp)) + self.minNp
        return fanDeviationNp


# 查询数据
aa = dbn.database_niu()  # 重命名
[conn, message] = aa.connection(
    db='ci_2', user='niuc',
    passwd='ni', host='')
# 链接数据库
print(message)
cur = conn.cursor()  # 获取一个游标
# 购买次数及金额
search = "SELECT * FROM niu"
dataNumMoney = aa.searchte(search, conn)
dataNumMoneyNp = np.array(dataNumMoney)  # 转成数组
dataNumMoneypd = DataFrame(dataNumMoneyNp, columns=['cishu', 'money'])  # 转成表格
# 次数百分比查看
dataNumpercent = dataNumMoneypd.groupby(dataNumMoneypd['cishu']).count()
dataNumnp = np.array(dataNumpercent.reset_index())


cishuXiao = 1
cishuDa = 60
jineDa = 1000000
jineXiao = 100000
dataNumMoneyNpCs = dataNumMoneyNp[(dataNumMoneyNp[:, 0] > cishuXiao) & (
    dataNumMoneyNp[:, 0] < cishuDa)]  # 投资次数处理
dataNumMoneyNpje = dataNumMoneyNpCs[
    (dataNumMoneyNpCs[:, 1] < jineDa) & (dataNumMoneyNpCs[:, 1] > jineXiao)]  # 投资金额处理

x = dataNumMoneyNpje[:, 0]
y = dataNumMoneyNpje[:, 1]
plt.scatter(x, y, label="training points")
plt.show()
# 原始构造
# estimator = KMeans(n_clusters=3)#构造聚类器
# estimator.fit(dataNumMoneyNp)  # 聚类
# label_pred = estimator.labels_  # 获取聚类标签
# centroids = estimator.cluster_centers_  # 获取聚类中心
# inertia = estimator.inertia_  # 获取聚类准则的总和

# 标准化（Z-Score）
mean = np.mean(dataNumMoneyNpje, axis=0).astype('float')  # 均值
std = np.std(dataNumMoneyNpje.astype('float'), axis=0)  # 方差
X_scaled = preprocessing.scale(dataNumMoneyNpje)

num_clusters = 6
km_cluster = KMeans(n_clusters=num_clusters, max_iter=3000, n_init=100,
                    init='k-means++', n_jobs=-1)

result = km_cluster.fit_predict(X_scaled)  # 分类标签

colors = ['teal', 'yellowgreen', 'gold', 'red', 'beige', 'sienna']
for k in range(0, num_clusters):
    xplt = dataNumMoneyNpje[(result == k), 0]
    yplt = dataNumMoneyNpje[result == k, 1]
    plt.scatter(xplt, yplt, color=colors[k], linewidth=2)
centroids = km_cluster.cluster_centers_  # 获取聚类中心
# 还原
out = centroids * std + mean
plt.scatter(out[:, 0], out[:, 1],
            marker='x', s=169, linewidths=3,
            color='b', zorder=10)
plt.show()


# 离差标准化
kkk = DeviationStandardization()
X_scaled = kkk.MaxMinNormalization(dataNumMoneyNpje)
num_clusters = 6
km_cluster1 = KMeans(n_clusters=num_clusters, max_iter=3000, n_init=100,
                     init='k-means++', n_jobs=-1)

result = km_cluster1.fit_predict(X_scaled)  # 分类标签

colors = ['teal', 'yellowgreen', 'gold', 'red', 'beige', 'sienna']
for k in range(0, num_clusters):
    xplt = dataNumMoneyNpje[(result == k), 0]
    yplt = dataNumMoneyNpje[result == k, 1]
    plt.scatter(xplt, yplt, color=colors[k], linewidth=2)
centroids1 = km_cluster1.cluster_centers_  # 获取聚类中心
z = kkk.fanization(centroids1)

plt.scatter(z[:, 0], z[:, 1],
            marker='x', s=169, linewidths=3,
            color='b', zorder=10)

plt.show()


# 文件存储
# 数据导出
outwb = Workbook()  # 新建文件
wo = outwb.active  # 获取激活文件，确认当前工作表
careerSheet1 = outwb.create_sheet('投资次数', 0)  # 插入一个sheet叫career，创建当前工作表
careerSheet1.append(['投资次数', '投资人次'])
careerSheet2 = outwb.create_sheet('投资金额', 1)  # 插入一个sheet叫career，创建当前工作表
careerSheet2.append(['投资次数', '投资金额'])
for x in np.array(dataNumnp):
    careerSheet1.append(list(x))
for x in np.array(dataNumMoneyNpCs):
    careerSheet2.append(list(x))
try:
    os.getcwd()
    os.chdir('e:\\data')
    name = '活动人数投资次数及金额聚类分析.xlsx'
    outwb.save(name)  # 只能文件有save属性
except Exception as err:
    fillte = '导出失败:' + str(err)
    print(fillte)
else:
    succefull = '导出成功'
    print(succefull)
